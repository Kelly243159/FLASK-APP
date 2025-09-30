from fasthtml.common import *  # FastHTML + Starlette + HTMX helpers
import pandas as pd
from datetime import datetime, timedelta
from unicodedata import normalize
import re, io, uuid, time, zipfile
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from starlette.responses import StreamingResponse, Response

# -------------------------------------------------
# Helpers de dados
# -------------------------------------------------
def _only_digits(s: str) -> str:
    return re.sub(r"\D+", "", str(s) if s is not None else "")

def _norm(s: str) -> str:
    s = "" if s is None else str(s)
    s = normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    return re.sub(r"\s+", " ", s.strip()).lower()

def _pick_col(df, candidates):
    m = {_norm(c): c for c in df.columns}
    for c in candidates:
        k = _norm(c)
        if k in m: return m[k]
    for c in candidates:
        k = re.sub(r"[^a-z0-9 ]+", "", _norm(c))
        for kk, orig in m.items():
            if re.sub(r"[^a-z0-9 ]+", "", kk) == k:
                return orig
    return None

def _status(venc_dt, today=None):
    if today is None: today = datetime.today()
    if pd.isna(venc_dt): return "Sem data"
    d = (venc_dt - today).days
    if d < 0: return "Vencido"
    if d <= 30: return "A vencer"
    return "No prazo"

def gerar_relatorio(df_sieg, df_cert):
    cnpj_sieg = _pick_col(df_sieg, ["CPF_CNPJ", "CNPJ", "CPF/CNPJ", "CPF CNPJ"])
    cnpj_cert = _pick_col(df_cert, ["CPF_CNPJ", "CNPJ", "CPF/CNPJ", "CPF CNPJ"])
    if not cnpj_sieg or not cnpj_cert:
        raise ValueError("Não encontrei a coluna CPF_CNPJ/CNPJ em uma das planilhas.")

    col_resp = _pick_col(df_sieg, ["Responsável", "Responsavel"])
    col_emp  = _pick_col(df_sieg, ["Empresa", "Razão Social", "Razao Social", "Cliente", "Nome do Cliente"])
    col_venc = _pick_col(df_cert, ["Vencimento", "Validade", "Data de Vencimento", "Data Vencimento"])

    df_sieg["_CPF_CNPJ_"] = df_sieg[cnpj_sieg].map(_only_digits)
    df_cert["_CPF_CNPJ_"] = df_cert[cnpj_cert].map(_only_digits)

    keep = ["_CPF_CNPJ_"] + ([col_venc] if col_venc else [])
    df_cert_small = df_cert[keep].copy()

    merged = pd.merge(df_sieg, df_cert_small, on="_CPF_CNPJ_", how="left")

    out = pd.DataFrame()
    out["Responsavel"] = merged[col_resp] if col_resp else ""
    out["Empresa"]     = merged[col_emp] if col_emp else ""
    out["CPF_CNPJ"]    = merged["_CPF_CNPJ_"]

    if col_venc:
        venc = pd.to_datetime(merged[col_venc], errors="coerce", dayfirst=True)
        out["Vencimento"] = venc.dt.strftime("%d/%m/%Y").fillna("")
        out["Status"]     = [_status(d) for d in venc]
    else:
        out["Vencimento"] = ""
        out["Status"]     = "Sem data"

    return out

def to_styled_html(df: pd.DataFrame) -> str:
    cols = list(df.columns)
    ths = ''.join(f"<th>{c}</th>" for c in cols)
    rows_html = []
    for _, row in df.iterrows():
        status = row.get("Status", "")
        if status == "Vencido":
            bg = "#FFC7CE"
        elif status == "A vencer":
            bg = "#FFEB9C"
        elif status == "No prazo":
            bg = "#C6EFCE"
        else:
            bg = "#ECEFF1"
        tds = []
        for c in cols:
            if c == "Status":
                tds.append(f'<td style="background:{bg};color:#000;font-weight:600">{row[c]}</td>')
            else:
                tds.append(f"<td>{row[c]}</td>")
        rows_html.append("<tr>" + "".join(tds) + "</tr>")
    tbody = "".join(rows_html)
    table = f"""
    <div class="table-wrap">
    <table class="data-table">
      <thead><tr>{ths}</tr></thead>
      <tbody>{tbody}</tbody>
    </table>
    </div>
    """
    return table

def make_excel_bytes(df: pd.DataFrame, sheet_name="Relatorio") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Cabeçalho
        header_fill = PatternFill(start_color="222222", end_color="222222", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[get_column_letter(col)].width = max(12, len(str(df.columns[col-1])) + 4)

        # Coloração da coluna Status
        status_col_idx = None
        for i, name in enumerate(df.columns, start=1):
            if name == "Status":
                status_col_idx = i
                break
        if status_col_idx:
            fill_vencido  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            fill_avencer  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            fill_noprazo  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fill_semdata  = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
            for r in range(2, len(df) + 2):
                c = ws.cell(row=r, column=status_col_idx)
                val = (c.value or "").strip()
                if val == "Vencido":
                    c.fill = fill_vencido
                elif val == "A vencer":
                    c.fill = fill_avencer
                elif val == "No prazo":
                    c.fill = fill_noprazo
                else:
                    c.fill = fill_semdata
    buf.seek(0)
    return buf.getvalue()

def make_zip_with_excel(xlsx_bytes: bytes, inner_name="Relatorio_SIEG_Certificados.xlsx") -> bytes:
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(inner_name, xlsx_bytes)
    zbuf.seek(0)
    return zbuf.getvalue()

# -------------------------------------------------
# Armazenamento com TTL (evita expiração imediata)
# -------------------------------------------------
class SafeStore:
    def __init__(self, ttl_seconds: int = 600):
        self.ttl = ttl_seconds
        self._data: dict[str, tuple[bytes, float, str]] = {}  # key -> (bytes, exp_ts, filename)

    def put(self, data: bytes, filename: str) -> str:
        key = str(uuid.uuid4())
        exp = time.time() + self.ttl
        self._data[key] = (data, exp, filename)
        return key

    def get(self, key: str):
        item = self._data.get(key)
        if not item:
            return None
        data, exp, filename = item
        # limpa expirados no acesso
        if time.time() > exp:
            try: del self._data[key]
            except: pass
            return None
        return data, filename

    def purge(self):
        now = time.time()
        for k, (_, exp, _) in list(self._data.items()):
            if now > exp:
                del self._data[k]

STORE = SafeStore(ttl_seconds=600)  # 10 minutos

# -------------------------------------------------
# FastHTML app (UI)
# -------------------------------------------------
app, rt = fast_app()
def global_css() -> FT:
    return Style("""
    :root{
      --brand:#0ea5e9; --brand-2:#1e3a8a;
      --glass-bg: rgba(255,255,255,0.06); --glass-br: 18px;
      --text-contrast:#eef7ff; --muted:#b9c8d6;
    }
    body{min-height:100vh;background:linear-gradient(135deg,#0b1020 0%,#0d1b2a 35%,#0a1a29 100%);color:var(--text-contrast);}
    .topbar{backdrop-filter:blur(8px);background:linear-gradient(90deg,rgba(14,165,233,0.18),rgba(30,58,138,0.18));border-bottom:1px solid rgba(255,255,255,0.06);}
    .brand{font-weight:800;letter-spacing:.3px;}
    .glass{background:var(--glass-bg);border:1px solid rgba(255,255,255,0.12);border-radius:var(--glass-br);box-shadow:0 10px 30px rgba(0,0,0,.25);padding:1.25rem;}
    label,.lbl{color:#e6f6ff;font-weight:700;letter-spacing:.2px;}
    small.hint{color:var(--muted);display:block;margin:.25rem 0 .5rem;}
    .filebox{appearance:none;width:100%;padding:1.1rem;background:rgba(255,255,255,0.08);border:2px dashed var(--brand);border-radius:14px;color:#e9f7ff;transition:.15s;}
    .filebox:hover{background:rgba(255,255,255,0.11);filter:brightness(1.03);border-color:#5fd3ff;}
    .filebox:focus{outline:none;box-shadow:0 0 0 4px rgba(14,165,233,.25);border-color:#7ee1ff;}
    .btn-primary{background:linear-gradient(90deg,var(--brand),#38bdf8);color:#001018;border:none;border-radius:12px;font-weight:800;padding:.7rem 1rem;display:inline-block;text-decoration:none;}
    .table-wrap{overflow:auto;max-height:70vh;border-radius:14px;border:1px solid rgba(255,255,255,0.12);}
    .data-table{width:100%;border-collapse:separate;border-spacing:0;background:rgba(255,255,255,0.02);color:#eaf6ff;}
    .data-table thead th{position:sticky;top:0;background:linear-gradient(90deg,rgba(14,165,233,0.32),rgba(30,58,138,0.32));backdrop-filter:blur(6px);color:#fff;font-weight:800;}
    .data-table th,.data-table td{padding:.7rem .85rem;}
    .footer{opacity:.85;font-size:.9rem;}
    """)

def hero() -> FT:
    return Header(
        Nav(
            Ul(Li(Strong("📊 SIEG x Certificados", cls="brand"))),
            Ul(Li(A("Ajuda", href="#", cls="contrast"))),
            cls="container"
        ),
        cls="topbar"
    )

def upload_form(msg: str | None = None) -> FT:
    alert = (Div(msg, role="alert", cls="container glass") if msg else None)
    return (
        global_css(),
        hero(),
        Main(
            Section(
                Article(
                    Hgroup(H1("Comparador por CNPJ"),
                           P("Anexe as planilhas SIEG e Certificados e gere o relatório com status de vencimento.")
                    ),
                    Form(enctype="multipart/form-data", hx_post=process_upload, hx_target="#result", hx_swap="innerHTML")(
                        Grid(
                            Fieldset(
                                Label("Planilha SIEG (xlsx/xls)", cls="lbl"),
                                Small("Clique para escolher ou arraste o arquivo aqui.", cls="hint"),
                                Input(type="file", name="file_sieg", accept=".xlsx,.xls", required=True, cls="filebox"),
                            ),
                            Fieldset(
                                Label("Planilha Certificados (xlsx/xls)", cls="lbl"),
                                Small("Aceita .xlsx e .xls — o CPF/CNPJ será normalizado.", cls="hint"),
                                Input(type="file", name="file_cert", accept=".xlsx,.xls", required=True, cls="filebox"),
                            ),
                        ),
                        Button("🚀 GERAR RELATÓRIO", type="submit", cls="btn-primary")
                    ),
                    cls="glass"
                ),
                cls="container"
            ),
            Section(id="result", cls="container", style="margin-top:1rem;"),
        ),
        Footer(P("MV Contabilidade • SGQ · ISO 9001", cls="footer container"))
    )

# ---------- Rotas ----------
app, rt = app, rt  # já definidos

@rt
def index():  # GET /
    return Titled("SIEG x Certificados", upload_form())

@rt
async def process_upload(request: Request, file_sieg: UploadFile, file_cert: UploadFile):
    try:
        df_sieg = pd.read_excel(io.BytesIO(await file_sieg.read()), dtype=str)
        df_cert = pd.read_excel(io.BytesIO(await file_cert.read()), dtype=str)
        resultado = gerar_relatorio(df_sieg, df_cert)

        html_table = to_styled_html(resultado)
        xbytes = make_excel_bytes(resultado)
        # nomes previsíveis ajudam alguns antivírus a não sinalizar:
        fname = f"Relatorio_SIEG_Certificados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        key = STORE.put(xbytes, filename=fname)

        base = f"{request.url.scheme}://{request.url.netloc}"
        url_xlsx = base + download_xlsx.to(key=key)
        url_zip  = base + download_zip.to(key=key)

        return Section(
            Article(
                H2("📋 Preview do Relatório"),
                Div(NotStr(html_table)),
                Hr(),
                H3("📥 Download"),
                P(A("Baixar Excel (.xlsx)", href=url_xlsx, cls="btn-primary", target="_blank"), " ",
                  A("Baixar ZIP (.zip)", href=url_zip, cls="btn-primary", style="margin-left:.5rem", target="_blank")),
                Small("Os links expiram em 10 minutos."),
                cls="glass"
            ),
        )
    except Exception as e:
        return Article(P("❌ Erro: " + str(e)), cls="glass")

@rt
def download_xlsx(key: str):
    got = STORE.get(key)
    if not got:
        return Article(P("⏲️ Link expirado. Gere o relatório novamente."), cls="glass container")
    data, filename = got
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Content-Type-Options": "nosniff",
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0",
    }
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@rt
def download_zip(key: str):
    got = STORE.get(key)
    if not got:
        return Article(P("⏲️ Link expirado. Gere o relatório novamente."), cls="glass container")
    data, filename = got
    zip_bytes = make_zip_with_excel(data, inner_name=filename)
    headers = {
        "Content-Disposition": f'attachment; filename="{filename.replace(".xlsx",".zip")}"',
        "X-Content-Type-Options": "nosniff",
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0",
    }
    return StreamingResponse(io.BytesIO(zip_bytes), media_type="application/zip", headers=headers)

# ---------- start ----------
serve()  # http://localhost:5001


